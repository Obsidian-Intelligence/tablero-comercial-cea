"""Genera Excel_de_Ventas_CEA.xlsx: la plantilla operativa del Tablero Comercial CEA.

Reproducible: correr `python scripts/generar_excel_ventas.py` desde la raíz del
repo regenera el archivo desde cero. Los nombres de hoja y columna DEBEN
coincidir con lo que leen src/sheets.py y src/etl.py — si cambias algo aquí,
cambia también esos dos archivos.
"""
from __future__ import annotations

import datetime as dt
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from src.constants import (  # noqa: E402
    APROBADORES,
    CANALES,
    ESTRATEGIAS,
    ETAPAS,
    KPIS_MANUALES,
    KR_IDS,
    MOTIVOS_PERDIDA,
    PLAZAS,
    PROGRAMAS,
    RESPONSABLES,
    TIPOS_DESCUENTO,
    VIGENCIA_DESCUENTO,
)

AUTOR_COMENTARIOS = "Alfonso"
SALIDA = REPO_ROOT / "Excel_de_Ventas_CEA.xlsx"

GRIS_CALCULADO = PatternFill("solid", fgColor="EAEAEA")
GRIS_EJEMPLO = PatternFill("solid", fgColor="D9D9D9")
VERDE_ALUMNO = PatternFill("solid", fgColor="C6E0B4")
GRIS_PERDIDO = PatternFill("solid", fgColor="D9D9D9")
ROJO_TOPE = PatternFill("solid", fgColor="F8CBAD")

FUENTE_ENCABEZADO = Font(bold=True)
FMT_FECHA = "DD/MM/AAAA"
FMT_MXN = '#,##0 "MXN"'


def nueva_hoja(wb: Workbook, nombre: str, *, con_grid: bool = True):
    ws = wb.create_sheet(nombre)
    if not con_grid:
        ws.sheet_view.showGridLines = False
    return ws


def encabezados(ws, columnas: list[str]) -> None:
    for i, nombre in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=i, value=nombre)
        celda.font = FUENTE_ENCABEZADO
    ws.freeze_panes = "A2"


def comentario(ws, celda: str, texto: str) -> None:
    c = Comment(texto, AUTOR_COMENTARIOS)
    c.width = 320
    c.height = 140
    ws[celda].comment = c


def agregar_validacion_lista(wb: Workbook, ws, columna_letra: str, nombre_rango: str, fila_inicio: int, fila_fin: int) -> None:
    dv = DataValidation(type="list", formula1=f"={nombre_rango}", allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"{columna_letra}{fila_inicio}:{columna_letra}{fila_fin}")


def agregar_validacion_fecha(ws, columna_letra: str, fila_inicio: int, fila_fin: int) -> None:
    dv = DataValidation(type="date", operator="greaterThan", formula1="DATE(2020,1,1)", allow_blank=True)
    dv.error = "Escribe una fecha válida (DD/MM/AAAA)."
    ws.add_data_validation(dv)
    dv.add(f"{columna_letra}{fila_inicio}:{columna_letra}{fila_fin}")
    for fila in range(fila_inicio, fila_fin + 1):
        ws[f"{columna_letra}{fila}"].number_format = FMT_FECHA


def agregar_validacion_decimal(ws, columna_letra: str, fila_inicio: int, fila_fin: int, minimo: float, maximo: float) -> None:
    dv = DataValidation(type="decimal", operator="between", formula1=str(minimo), formula2=str(maximo), allow_blank=True)
    dv.error = f"Debe estar entre {minimo} y {maximo}."
    ws.add_data_validation(dv)
    dv.add(f"{columna_letra}{fila_inicio}:{columna_letra}{fila_fin}")


def agregar_validacion_entero_positivo(ws, columna_letra: str, fila_inicio: int, fila_fin: int) -> None:
    dv = DataValidation(type="whole", operator="greaterThanOrEqual", formula1="0", allow_blank=True)
    dv.error = "Debe ser un número entero de 0 en adelante."
    ws.add_data_validation(dv)
    dv.add(f"{columna_letra}{fila_inicio}:{columna_letra}{fila_fin}")


def pintar_filas_ejemplo(ws, fila_inicio: int, fila_fin: int, num_columnas: int) -> None:
    for fila in range(fila_inicio, fila_fin + 1):
        for col in range(1, num_columnas + 1):
            ws.cell(row=fila, column=col).fill = GRIS_EJEMPLO


def congelar_y_filtrar(ws, num_columnas: int) -> None:
    ultima_col = get_column_letter(num_columnas)
    ws.auto_filter.ref = f"A1:{ultima_col}1"


def ajustar_anchos(ws, anchos: dict[str, int]) -> None:
    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho


def construir_config(wb: Workbook) -> dict[str, tuple[str, int]]:
    """Crea la hoja Config con una columna por lista, y un rango con nombre
    por cada una (usado por las validaciones de lista de las demás hojas).
    Regresa {nombre_lista: (columna_letra, cantidad_de_valores)}."""
    ws = nueva_hoja(wb, "Config")
    comentario(ws, "A1", "Solo Alfonso edita esta hoja. Agregar aquí = aparece en las listas de todas las hojas.")

    listas = {
        "etapas": ETAPAS,
        "canales": CANALES,
        "plazas": PLAZAS,
        "programas": PROGRAMAS,
        "responsables": RESPONSABLES,
        "aprobadores": APROBADORES,
        "tipos_descuento": TIPOS_DESCUENTO,
        "motivos_perdida": MOTIVOS_PERDIDA,
        "vigencia_descuento": VIGENCIA_DESCUENTO,
        "estrategias": ESTRATEGIAS + ["Total"],
        "kr_ids": KR_IDS,
        "kpis_manuales": KPIS_MANUALES,
        "si_no_parcial": ["Sí", "No", "Parcial"],
    }

    rangos: dict[str, tuple[str, int]] = {}
    for i, (nombre_lista, valores) in enumerate(listas.items(), start=1):
        col_letra = get_column_letter(i)
        ws.cell(row=1, column=i, value=nombre_lista).font = FUENTE_ENCABEZADO
        for j, valor in enumerate(valores, start=2):
            ws.cell(row=j, column=i, value=valor)
        rango_nombre = f"lista_{nombre_lista}"
        wb.defined_names[rango_nombre] = DefinedName(
            rango_nombre, attr_text=f"'Config'!${col_letra}$2:${col_letra}${len(valores) + 1}"
        )
        rangos[nombre_lista] = (col_letra, len(valores))

    # Tabla programa -> mensualidad plena, para el formato condicional de Descuentos.
    with open(REPO_ROOT / "data" / "precios_cea.json", encoding="utf-8") as f:
        precios = json.load(f)["programas"]
    mensualidad_por_programa = {
        "Lic. Administración": next(p["mensualidad_plena"] for p in precios if p["programa"].startswith("Licenciatura")),
        "Lic. Contaduría": next(p["mensualidad_plena"] for p in precios if p["programa"].startswith("Licenciatura")),
        "Lic. Derecho": next(p["mensualidad_plena"] for p in precios if p["programa"].startswith("Licenciatura")),
        "Bachillerato": next(p["mensualidad_plena"] for p in precios if p["programa"] == "Bachillerato Sabatino" and p["plaza"] == "Puebla"),
        "Maestría Derecho Penal": next(p["mensualidad_plena"] for p in precios if p["programa"].startswith("Maestría")),
    }
    col_prog = get_column_letter(len(listas) + 1)
    col_mens = get_column_letter(len(listas) + 2)
    ws.cell(row=1, column=len(listas) + 1, value="programa_precio").font = FUENTE_ENCABEZADO
    ws.cell(row=1, column=len(listas) + 2, value="mensualidad_plena").font = FUENTE_ENCABEZADO
    for j, (programa, monto) in enumerate(mensualidad_por_programa.items(), start=2):
        ws.cell(row=j, column=len(listas) + 1, value=programa)
        ws.cell(row=j, column=len(listas) + 2, value=monto)
    wb.defined_names["tabla_mensualidad_plena"] = DefinedName(
        "tabla_mensualidad_plena",
        attr_text=f"'Config'!${col_prog}$2:${col_mens}${len(mensualidad_por_programa) + 1}",
    )

    ajustar_anchos(ws, {get_column_letter(i): 22 for i in range(1, len(listas) + 3)})
    ws.freeze_panes = "A2"
    congelar_y_filtrar(ws, len(listas) + 2)
    return rangos


def construir_prospectos(wb: Workbook, rangos: dict[str, tuple[str, int]]) -> None:
    columnas = ["id", "nombre_prospecto", "fecha_registro", "plaza", "programa", "canal_origen", "referido_por",
                "etapa", "motivo_perdida", "responsable", "telefono", "telefono_alt", "email", "notas",
                "intentos_contacto"]
    ws = nueva_hoja(wb, "Prospectos")
    encabezados(ws, columnas)

    comentario(ws, "A1", "NO editar ni reciclar. Siguiente número libre. Es la huella del prospecto en todo el sistema.")
    comentario(ws, "B1", "Obligatorio: sin nombre no hay a quién contactar ni de quién hablar en el comité. "
                         "Nunca sale de este Excel — el tablero público la descarta antes de cualquier gráfica.")
    comentario(ws, "C1", "Día en que la persona preguntó por primera vez.")
    comentario(ws, "D1", "Sede donde quiere estudiar, no donde preguntó.")
    comentario(ws, "F1", "¿Cómo nos encontró? Si te dice 'me recomendó X', es E4 Referidos y llenas la col. G. "
                         "Esta columna decide dónde invertimos: no adivines, pregunta.")
    comentario(ws, "G1", "Solo si canal = E4. Sin esto no se paga el premio.")
    comentario(ws, "H1", "Actualízala EN EL MOMENTO. Cada cambio queda registrado automáticamente con fecha (hoja Eventos).")
    comentario(ws, "I1", "Solo si etapa = Perdido. Sé honesta: 'Precio' nos sirve más que 'Otro'.")
    comentario(ws, "K1", "Solo para seguimiento. JAMÁS va al tablero.")
    comentario(ws, "L1", "Teléfono de un familiar o conocido, por si no contesta el suyo — sube la tasa de "
                         "contacto. Igual que el teléfono principal, nunca sale de este Excel.")
    comentario(ws, "M1", "Opcional pero recomendado: habilita campañas futuras de WhatsApp/email. Nunca sale de este Excel.")
    comentario(ws, "N1", "Contexto útil: horario que le acomoda, empresa donde trabaja, etc.")
    comentario(ws, "O1", "Cuántas veces lo buscamos antes de responder o darlo por perdido. Mínimo 3 intentos "
                         "antes de marcar 'No responde'.")

    fila_inicio, fila_fin = 2, 4  # 3 filas EJEMPLO
    ejemplos = [
        ["EJEMPLO — P-0001", "EJEMPLO — Nombre Apellido", dt.date(2026, 7, 1), "Puebla", "Lic. Administración", "E6", "",
         "Contactado", "", "Seguimiento a prospectos", "2221234567", "", "", "Prefiere sábados", 1],
        ["EJEMPLO — P-0002", "EJEMPLO — Nombre Apellido", dt.date(2026, 7, 2), "Huauchinango", "Bachillerato", "E4", "P-0001",
         "Inscripción provisional", "", "Alfonso", "2227654321", "2227654322", "ejemplo@correo.com", "Refirió su hermano", 2],
        ["EJEMPLO — P-0003", "EJEMPLO — Nombre Apellido", dt.date(2026, 6, 20), "Zacatlán", "Lic. Derecho", "E6", "",
         "Perdido", "Precio", "Control Escolar", "2229998888", "", "", "", 4],
    ]
    for i, fila_valores in enumerate(ejemplos):
        for j, valor in enumerate(fila_valores, start=1):
            ws.cell(row=fila_inicio + i, column=j, value=valor)
    pintar_filas_ejemplo(ws, fila_inicio, fila_fin, len(columnas))

    fila_datos_fin = 500  # rango de captura + validaciones
    agregar_validacion_fecha(ws, "C", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "D", "lista_plazas", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "E", "lista_programas", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "F", "lista_canales", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "H", "lista_etapas", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "I", "lista_motivos_perdida", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "J", "lista_responsables", fila_inicio, fila_datos_fin)
    agregar_validacion_entero_positivo(ws, "O", fila_inicio, fila_datos_fin)

    # Formato condicional: etapa=Alumno -> verde, Perdido -> gris.
    from openpyxl.formatting.rule import FormulaRule
    rango_completo = f"A{fila_inicio}:O{fila_datos_fin}"
    ws.conditional_formatting.add(rango_completo, FormulaRule(formula=[f"$H{fila_inicio}=\"Alumno\""], fill=VERDE_ALUMNO))
    ws.conditional_formatting.add(rango_completo, FormulaRule(formula=[f"$H{fila_inicio}=\"Perdido\""], fill=GRIS_PERDIDO))

    congelar_y_filtrar(ws, len(columnas))
    ajustar_anchos(ws, {"A": 16, "B": 22, "C": 14, "D": 14, "E": 20, "F": 12, "G": 14, "H": 20, "I": 16,
                         "J": 14, "K": 14, "L": 14, "M": 22, "N": 30, "O": 16})


def construir_eventos(wb: Workbook) -> None:
    columnas = ["id_prospecto", "etapa_anterior", "etapa_nueva", "fecha_hora", "usuario"]
    ws = nueva_hoja(wb, "Eventos")
    encabezados(ws, columnas)
    comentario(ws, "A1", "Historial automático. Si escribes aquí, rompes las métricas de velocidad del embudo.")
    congelar_y_filtrar(ws, len(columnas))
    ajustar_anchos(ws, {"A": 16, "B": 20, "C": 20, "D": 20, "E": 20})


def construir_instrucciones(wb: Workbook) -> None:
    ws = nueva_hoja(wb, "INSTRUCCIONES", con_grid=False)
    ws.column_dimensions["A"].width = 100

    lineas = [
        ("titulo", "Excel de Ventas CEA — cómo se llena"),
        ("texto", ""),
        ("sub", "¿Qué es esto?"),
        ("texto", "Este archivo alimenta el Tablero Comercial CEA que se revisa cada lunes. "
                   "Cada hoja corresponde a una parte del ritual semanal: prospección, avance de "
                   "metas, alumnos activos, descuentos y los KPIs que no salen de ningún otro lado."),
        ("texto", ""),
        ("sub", "Regla de oro"),
        ("texto", "Si no está aquí, no existió. El tablero solo cree lo que esté capturado antes "
                   "del lunes 8:00 am."),
        ("texto", ""),
        ("sub", "Quién llena qué"),
        ("texto", "• Asesoría comercial → hoja Prospectos."),
        ("texto", "• Control Escolar → hoja Alumnos_Activos y parte de KPI_Manual (NPS, bajas)."),
        ("texto", "• Administración y Finanzas → hoja Alumnos_Activos (ingreso/facturable/cobrado), Descuentos_Otorgados y cobranza."),
        ("texto", "• Mercadotecnia → KPI_Manual (reseñas, prospectos por canal de pauta)."),
        ("texto", "• Alfonso → hojas Metas y OKR_Avance, en vivo durante el comité de los lunes."),
        ("texto", ""),
        ("sub", "Los 3 errores que rompen el tablero"),
        ("texto", "1. Escribir un valor a mano en vez de elegir de la lista desplegable."),
        ("texto", "2. Borrar o reciclar el id de un prospecto."),
        ("texto", "3. Capturar el teléfono en la columna de notas."),
        ("texto", ""),
        ("sub", "Nunca tocar la hoja Eventos"),
        ("texto", "Se llena sola con la macro VBA instalada en este archivo (ver vba/registro_eventos.bas "
                   "para el código y los pasos de instalación). Escribir ahí a mano rompe las métricas de "
                   "velocidad del embudo. La macro solo funciona si el archivo se abre en Excel de escritorio "
                   "con macros habilitadas — si Excel pide 'Habilitar contenido' al abrir, hay que aceptarlo."),
        ("texto", ""),
        ("sub", "Aviso de privacidad"),
        ("texto", "Los datos de contacto (teléfono, notas) solo se usan para dar seguimiento a la "
                   "inscripción, conforme a la LFPDPPP. El tablero público nunca muestra nombres ni "
                   "teléfonos — esas columnas se descartan antes de llegar a cualquier gráfica."),
        ("texto", ""),
        ("sub", "Antes de arrancar"),
        ("texto", "Borra las 3 filas que empiezan con \"EJEMPLO —\" en cada hoja de captura. Son "
                   "solo para mostrar cómo se llena, no son datos reales."),
    ]

    fila = 1
    for tipo, texto in lineas:
        celda = ws.cell(row=fila, column=1, value=texto)
        if tipo == "titulo":
            celda.font = Font(bold=True, size=16)
        elif tipo == "sub":
            celda.font = Font(bold=True, size=12)
        else:
            celda.alignment = Alignment(wrap_text=True, vertical="top")
        fila += 1


def construir_metas(wb: Workbook, rangos: dict[str, tuple[str, int]]) -> None:
    columnas = ["semana_iso", "estrategia", "meta_alumnos", "meta_prospectos", "responsable", "comentario_cierre"]
    ws = nueva_hoja(wb, "Metas")
    encabezados(ws, columnas)
    comentario(ws, "F1", "Se llena al cierre: si no llegamos, ¿por qué? Una frase honesta. Esto se lee LITERAL "
                         "en el tablero y en el comité.")

    fila_inicio, fila_fin = 2, 4
    ejemplos = [
        ["EJEMPLO — 2026-W29", "E4", 2, 8, "Alfonso", ""],
        ["EJEMPLO — 2026-W29", "E6", 1, 6, "Seguimiento a prospectos", "Se retrasó la campaña por ajustes de presupuesto."],
        ["EJEMPLO — 2026-W29", "Total", 10, 42, "CEA Dirección", ""],
    ]
    for i, fila_valores in enumerate(ejemplos):
        for j, valor in enumerate(fila_valores, start=1):
            ws.cell(row=fila_inicio + i, column=j, value=valor)
    pintar_filas_ejemplo(ws, fila_inicio, fila_fin, len(columnas))

    fila_datos_fin = 500
    agregar_validacion_lista(wb, ws, "B", "lista_estrategias", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "E", "lista_responsables", fila_inicio, fila_datos_fin)

    congelar_y_filtrar(ws, len(columnas))
    ajustar_anchos(ws, {"A": 16, "B": 10, "C": 14, "D": 16, "E": 14, "F": 50})


def construir_alumnos_activos(wb: Workbook, rangos: dict[str, tuple[str, int]]) -> None:
    columnas = ["fecha_corte", "plaza", "programa", "total_activos", "ingreso_mensual_real_mxn",
                "facturable_mxn", "cobrado_mxn"]
    ws = nueva_hoja(wb, "Alumnos_Activos")
    encabezados(ws, columnas)
    comentario(ws, "D1", "Cuenta de alumnos vigentes al viernes. 5 minutos, una fila por plaza-programa.")
    comentario(ws, "F1", "Lo que DEBERÍA entrar este mes según lista de precios y descuentos vigentes.")
    comentario(ws, "G1", "Lo que SÍ entró. cobrado ÷ facturable = KPI 10 de cobranza (meta ≥90%).")

    fila_inicio, fila_fin = 2, 4
    ejemplos = [
        ["EJEMPLO — 2026-07-10", "Puebla", "Lic. Administración", 40, 64000, 68000, 61200],
        ["EJEMPLO — 2026-07-10", "Huauchinango", "Bachillerato", 25, 18000, 19000, 17500],
        ["EJEMPLO — 2026-07-10", "Zacatlán", "Bachillerato", 15, 9000, 9500, 8500],
    ]
    for i, fila_valores in enumerate(ejemplos):
        for j, valor in enumerate(fila_valores, start=1):
            ws.cell(row=fila_inicio + i, column=j, value=valor)
    pintar_filas_ejemplo(ws, fila_inicio, fila_fin, len(columnas))
    for fila in range(fila_inicio, fila_fin + 1):
        for col_letra in ("E", "F", "G"):
            ws[f"{col_letra}{fila}"].number_format = FMT_MXN

    fila_datos_fin = 500
    agregar_validacion_fecha(ws, "A", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "B", "lista_plazas", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "C", "lista_programas", fila_inicio, fila_datos_fin)
    for fila in range(fila_inicio, fila_datos_fin + 1):
        for col_letra in ("E", "F", "G"):
            ws[f"{col_letra}{fila}"].number_format = FMT_MXN

    congelar_y_filtrar(ws, len(columnas))
    ajustar_anchos(ws, {"A": 16, "B": 14, "C": 20, "D": 14, "E": 20, "F": 16, "G": 16})


def construir_descuentos(wb: Workbook, rangos: dict[str, tuple[str, int]]) -> None:
    columnas = ["fecha", "folio_alumno", "plaza", "programa", "estrategia_asociada", "tipo",
                "monto_mensual_mxn", "vigencia", "aprobado_por", "evidencia"]
    ws = nueva_hoja(wb, "Descuentos_Otorgados")
    encabezados(ws, columnas)
    comentario(ws, "A1", "Todo descuento vive aquí o no existe. Tope por alumno: 20% del valor anual. "
                         "Nada se acumula. Promociones vencen 28-ago-2026.")
    comentario(ws, "B1", "Es el folio del alumno (id P-xxxx), nunca su nombre. Búscalo en la hoja Prospectos.")

    fila_inicio, fila_fin = 2, 4
    ejemplos = [
        [dt.date(2026, 7, 5), "EJEMPLO — P-0002", "Huauchinango", "Bachillerato", "E4",
         "Inscripción 50% referido", 150, "Única vez", "Administración y Finanzas", "folio-102"],
        [dt.date(2026, 7, 8), "EJEMPLO — P-0015", "Puebla", "Lic. Contaduría", "E3",
         "Beca %", 180, "Permanente", "CEA Dirección", "correo-2026-07-08"],
        [dt.date(2026, 7, 10), "EJEMPLO — P-0022", "Zacatlán", "Bachillerato", "E9",
         "Inscripción gratis campaña", 0, "Única vez", "Administración y Finanzas", "acuerdo-tianguis-03"],
    ]
    for i, fila_valores in enumerate(ejemplos):
        for j, valor in enumerate(fila_valores, start=1):
            ws.cell(row=fila_inicio + i, column=j, value=valor)
    pintar_filas_ejemplo(ws, fila_inicio, fila_fin, len(columnas))

    fila_datos_fin = 500
    agregar_validacion_fecha(ws, "A", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "C", "lista_plazas", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "D", "lista_programas", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "E", "lista_estrategias", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "F", "lista_tipos_descuento", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "H", "lista_vigencia_descuento", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "I", "lista_aprobadores", fila_inicio, fila_datos_fin)
    for fila in range(fila_inicio, fila_datos_fin + 1):
        ws[f"G{fila}"].number_format = FMT_MXN

    # Formato condicional: monto_mensual_mxn > 20% de la mensualidad plena del programa -> rojo.
    from openpyxl.formatting.rule import FormulaRule
    rango_completo = f"A{fila_inicio}:J{fila_datos_fin}"
    formula_tope = (
        f'$G{fila_inicio}>0.2*IFERROR(VLOOKUP($D{fila_inicio},tabla_mensualidad_plena,2,FALSE),9999999)'
    )
    ws.conditional_formatting.add(rango_completo, FormulaRule(formula=[formula_tope], fill=ROJO_TOPE))

    congelar_y_filtrar(ws, len(columnas))
    ajustar_anchos(ws, {"A": 14, "B": 18, "C": 14, "D": 20, "E": 14, "F": 24, "G": 16, "H": 14, "I": 14, "J": 20})


def construir_kpi_manual(wb: Workbook, rangos: dict[str, tuple[str, int]]) -> None:
    columnas = ["semana_iso", "kpi", "plaza", "valor", "capturado_por"]
    ws = nueva_hoja(wb, "KPI_Manual")
    encabezados(ws, columnas)
    comentario(ws, "A1", "Solo estos 9 se capturan a mano; todo lo demás lo calcula el tablero. "
                         "Viernes antes de las 6 pm.")

    fila_inicio, fila_fin = 2, 4
    ejemplos = [
        ["EJEMPLO — 2026-W28", "NPS", "General", 52, "Control Escolar"],
        ["EJEMPLO — 2026-W28", "Citas agendadas", "Puebla", 9, "Seguimiento a prospectos"],
        ["EJEMPLO — 2026-W28", "Gasto pauta MXN", "General", 7200, "Mercadotecnia"],
    ]
    for i, fila_valores in enumerate(ejemplos):
        for j, valor in enumerate(fila_valores, start=1):
            ws.cell(row=fila_inicio + i, column=j, value=valor)
    pintar_filas_ejemplo(ws, fila_inicio, fila_fin, len(columnas))

    fila_datos_fin = 500
    agregar_validacion_lista(wb, ws, "B", "lista_kpis_manuales", fila_inicio, fila_datos_fin)

    # KPI_Manual.plaza acepta "General" además de las 3 plazas — eso no vive en
    # ninguna lista de Config todavía, así que se agrega una columna propia
    # ("plazas_o_general") en una columna nueva al final de Config.
    ws_config = wb["Config"]
    col_nueva = ws_config.max_column + 1
    col_letra = get_column_letter(col_nueva)
    ws_config.cell(row=1, column=col_nueva, value="plazas_o_general").font = FUENTE_ENCABEZADO
    valores_plaza_general = PLAZAS + ["General"]
    for i, valor in enumerate(valores_plaza_general, start=2):
        ws_config.cell(row=i, column=col_nueva, value=valor)
    wb.defined_names["lista_plazas_o_general"] = DefinedName(
        "lista_plazas_o_general",
        attr_text=f"'Config'!${col_letra}$2:${col_letra}${len(valores_plaza_general) + 1}",
    )
    congelar_y_filtrar(ws_config, col_nueva)  # extiende el auto_filter para cubrir la columna nueva

    agregar_validacion_lista(wb, ws, "C", "lista_plazas_o_general", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "E", "lista_responsables", fila_inicio, fila_datos_fin)

    congelar_y_filtrar(ws, len(columnas))
    ajustar_anchos(ws, {"A": 16, "B": 26, "C": 16, "D": 12, "E": 16})


def construir_okr_avance(wb: Workbook, rangos: dict[str, tuple[str, int]]) -> None:
    columnas = ["kr_id", "avance", "fecha_actualizacion", "comentario"]
    ws = nueva_hoja(wb, "OKR_Avance")
    encabezados(ws, columnas)
    comentario(ws, "B1", "0.0 = nada · 0.5 = a medio camino · 0.7 = éxito · 1.0 sistemático = la meta "
                         "estaba floja. Se actualiza en vivo cada lunes.")

    fila_inicio, fila_fin = 2, 4
    ejemplos = [
        ["EJEMPLO — KR1.2", 0.35, dt.date(2026, 7, 13), "Vamos a ritmo para llegar a 145 en septiembre."],
        ["EJEMPLO — KR2.5", 0.20, dt.date(2026, 7, 13), "Falta arrancar la encuesta NPS formal."],
        ["EJEMPLO — KR4.1", 0.60, dt.date(2026, 7, 13), "Cobranza mejorando pero aún bajo 90%."],
    ]
    for i, fila_valores in enumerate(ejemplos):
        for j, valor in enumerate(fila_valores, start=1):
            ws.cell(row=fila_inicio + i, column=j, value=valor)
    pintar_filas_ejemplo(ws, fila_inicio, fila_fin, len(columnas))

    fila_datos_fin = 500
    agregar_validacion_lista(wb, ws, "A", "lista_kr_ids", fila_inicio, fila_datos_fin)
    agregar_validacion_decimal(ws, "B", fila_inicio, fila_datos_fin, 0.0, 1.0)
    agregar_validacion_fecha(ws, "C", fila_inicio, fila_datos_fin)

    congelar_y_filtrar(ws, len(columnas))
    ajustar_anchos(ws, {"A": 12, "B": 12, "C": 18, "D": 50})


def _cargar_okrs() -> dict:
    with open(REPO_ROOT / "data" / "okrs.json", encoding="utf-8") as f:
        return json.load(f)


def construir_referencia_krs(wb: Workbook) -> None:
    """Hoja de solo lectura: qué significa cada KR, para consultar mientras se
    llena OKR_Avance sin tener que abrir el tablero. Se regenera desde
    data/okrs.json — si cambian los OKRs del trimestre, cambia ese archivo,
    no esta función."""
    okrs = _cargar_okrs()
    columnas = ["objetivo_id", "objetivo_titulo", "kr_id", "kr_texto", "marca"]
    ws = nueva_hoja(wb, "Referencia_KRs")
    encabezados(ws, columnas)
    comentario(ws, "A1", "Solo lectura, no se captura nada aquí. Consulta rápida de qué significa cada KR "
                         "antes de anotar su avance en OKR_Avance. ✅ confirmado · 📊 inferencia · 🎯 hipótesis.")

    fila = 2
    for objetivo in okrs["objetivos_q3"]:
        for kr in objetivo["krs"]:
            ws.cell(row=fila, column=1, value=objetivo["id"])
            ws.cell(row=fila, column=2, value=objetivo["titulo"])
            ws.cell(row=fila, column=3, value=kr["id"])
            ws.cell(row=fila, column=4, value=kr["texto"])
            ws.cell(row=fila, column=5, value=kr.get("marca", ""))
            ws.cell(row=fila, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            fila += 1

    congelar_y_filtrar(ws, len(columnas))
    ajustar_anchos(ws, {"A": 12, "B": 40, "C": 10, "D": 70, "E": 10})


def construir_referencia_kpis(wb: Workbook) -> None:
    """Hoja de solo lectura: los 12 KPIs permanentes del tablero (fuente,
    frecuencia, dueño, meta y contramétrica). De aquí salen los 9 que se
    capturan a mano en KPI_Manual. Se regenera desde data/okrs.json."""
    okrs = _cargar_okrs()
    columnas = ["n", "kpi", "fuente", "frecuencia", "dueno", "meta", "contrametrica"]
    ws = nueva_hoja(wb, "Referencia_KPIs")
    encabezados(ws, columnas)
    comentario(ws, "A1", "Solo lectura, no se captura nada aquí. Los 12 signos vitales del war room de los "
                         "lunes — la 'contrametrica' es para no optimizar el KPI a costa de otra cosa.")

    for i, kpi in enumerate(okrs["kpis_permanentes"], start=2):
        ws.cell(row=i, column=1, value=kpi["n"])
        ws.cell(row=i, column=2, value=kpi["kpi"])
        ws.cell(row=i, column=3, value=kpi["fuente"])
        ws.cell(row=i, column=4, value=kpi["frecuencia"])
        ws.cell(row=i, column=5, value=kpi["dueno"])
        ws.cell(row=i, column=6, value=kpi["meta"])
        ws.cell(row=i, column=7, value=kpi["contra"])
        ws.cell(row=i, column=7).alignment = Alignment(wrap_text=True, vertical="top")

    congelar_y_filtrar(ws, len(columnas))
    ajustar_anchos(ws, {"A": 6, "B": 32, "C": 20, "D": 12, "E": 16, "F": 22, "G": 45})


def construir_comite_lunes(wb: Workbook, rangos: dict[str, tuple[str, int]]) -> None:
    columnas = ["fecha", "semana_iso", "asistentes", "alumnos_activos_hoy", "kpis_en_rojo",
                "krs_movidos", "decision_de_la_semana", "responsable_decision", "fecha_limite",
                "decision_semana_pasada_cumplida", "nota_para_direccion"]
    ws = nueva_hoja(wb, "Comité_Lunes")
    encabezados(ws, columnas)
    comentario(ws, "A1", "30 minutos: 1) ¿cumplimos la decisión pasada? 2) tablero: alumnos, rojos, "
                         "KRs 3) UNA decisión nueva con dueño y fecha. Este historial es la memoria de "
                         "la operación.")
    comentario(ws, "D1", "Se lee del tablero, se anota aquí como acta.")
    comentario(ws, "E1", "Máximo 3, los peores. Del tablero.")
    comentario(ws, "F1", "¿Qué KR subió o bajó esta semana y por qué (qué dato lo movió)?")
    comentario(ws, "G1", "UNA sola decisión ejecutable. Regla del comité: sin decisión no hay sesión.")
    comentario(ws, "J1", "Se revisa ANTES de tomar la nueva.")
    comentario(ws, "K1", "Una frase para CEA Dirección si hay algo que necesite su decisión o su cartera.")

    fila_inicio, fila_fin = 2, 4
    ejemplos = [
        ["EJEMPLO — 06/07/2026", "2026-W28", "CEA Dirección, Administración y Finanzas, Mercadotecnia, Alfonso", 108,
         "Cobranza bajo 90%, CAC de E6 sobre umbral", "KR1.2 bajó por semana corta de puente.",
         "Confirmar convenio COPARMEX antes del viernes.", "CEA Dirección", dt.date(2026, 7, 10),
         "Sí", ""],
        ["EJEMPLO — 13/07/2026", "2026-W29", "CEA Dirección, Administración y Finanzas, Alfonso", 110,
         "Mediana 1ª respuesta sobre 5 min", "KR3.1 sin cambio, seguimos sin sostenerlo 4 semanas.",
         "Asesoría comercial revisa WhatsApp cada 2 horas fijas.", "Seguimiento a prospectos", dt.date(2026, 7, 17),
         "Parcial", ""],
        ["EJEMPLO — 20/07/2026", "2026-W30", "CEA Dirección, Administración y Finanzas, Mercadotecnia", 111,
         "Bajas del mes sobre meta", "KR2.1 empeoró; se investigan las 2 bajas de julio.",
         "Control Escolar llama a las 2 bajas para entender motivo real.", "Control Escolar",
         dt.date(2026, 7, 24), "No", "Necesitamos que CEA Dirección apruebe presupuesto de retención."],
    ]
    for i, fila_valores in enumerate(ejemplos):
        for j, valor in enumerate(fila_valores, start=1):
            ws.cell(row=fila_inicio + i, column=j, value=valor)
    pintar_filas_ejemplo(ws, fila_inicio, fila_fin, len(columnas))

    fila_datos_fin = 260  # ~5 años de lunes es de sobra
    agregar_validacion_fecha(ws, "A", fila_inicio, fila_datos_fin)
    agregar_validacion_fecha(ws, "I", fila_inicio, fila_datos_fin)
    agregar_validacion_lista(wb, ws, "H", "lista_responsables", fila_inicio, fila_datos_fin)

    agregar_validacion_lista(wb, ws, "J", "lista_si_no_parcial", fila_inicio, fila_datos_fin)

    congelar_y_filtrar(ws, len(columnas))
    ajustar_anchos(ws, {"A": 16, "B": 14, "C": 30, "D": 16, "E": 30, "F": 40, "G": 40, "H": 16, "I": 14, "J": 20, "K": 40})


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)

    construir_instrucciones(wb)
    rangos = construir_config(wb)  # Config debe existir antes que cualquier hoja con listas
    construir_prospectos(wb, rangos)
    construir_eventos(wb)
    construir_metas(wb, rangos)
    construir_alumnos_activos(wb, rangos)
    construir_descuentos(wb, rangos)
    construir_kpi_manual(wb, rangos)
    construir_okr_avance(wb, rangos)
    construir_referencia_krs(wb)
    construir_referencia_kpis(wb)
    construir_comite_lunes(wb, rangos)

    wb.active = 0
    wb.save(SALIDA)
    print(f"Escrito {SALIDA} con hojas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
