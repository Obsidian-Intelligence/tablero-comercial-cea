"""Verificación estructural de Excel_de_Ventas_CEA.xlsx — corre después de
generar_excel_ventas.py y confirma que el archivo cumple el checklist del spec."""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from src import etl  # noqa: E402

ARCHIVO = REPO_ROOT / "Excel_de_Ventas_CEA.xlsx"

HOJAS_ESPERADAS = [
    "INSTRUCCIONES", "Config", "Prospectos", "Eventos", "Metas", "Alumnos_Activos",
    "Descuentos_Otorgados", "KPI_Manual", "OKR_Avance", "Comité_Lunes",
]

HOJAS_CON_EJEMPLOS = [
    "Prospectos", "Metas", "Alumnos_Activos", "Descuentos_Otorgados", "KPI_Manual", "OKR_Avance", "Comité_Lunes",
]


def verificar() -> list[str]:
    errores = []
    wb = openpyxl.load_workbook(ARCHIVO)

    for hoja in HOJAS_ESPERADAS:
        if hoja not in wb.sheetnames:
            errores.append(f"Falta la hoja {hoja}")
    if errores:
        return errores  # sin hojas no tiene caso seguir

    # Columnas esperadas: la fuente de verdad es el propio ETL (lo que el
    # tablero realmente lee), no una lista aparte que se puede desincronizar.
    hojas_demo = etl.generar_datos_demo()
    hojas_con_columnas_fijas = [h for h in HOJAS_ESPERADAS if h not in ("INSTRUCCIONES", "Config")]
    for hoja in hojas_con_columnas_fijas:
        columnas_esperadas = list(hojas_demo[hoja].columns)
        columnas_reales = [c.value for c in wb[hoja][1]]
        if columnas_reales != columnas_esperadas:
            errores.append(
                f"{hoja}: encabezados no coinciden con src/etl.py — "
                f"esperado {columnas_esperadas}, encontrado {columnas_reales}"
            )

    for hoja in wb.sheetnames:
        ws = wb[hoja]
        if hoja == "INSTRUCCIONES":
            continue
        if ws.freeze_panes != "A2":
            errores.append(f"{hoja}: fila 1 no está congelada")
        if hoja != "INSTRUCCIONES" and not ws.auto_filter.ref:
            errores.append(f"{hoja}: no tiene auto_filter")
        primera_fila = [c.value for c in ws[1]]
        if not any(c.comment for c in ws[1]):
            errores.append(f"{hoja}: ningún encabezado tiene comentario")

    for hoja in HOJAS_CON_EJEMPLOS:
        ws = wb[hoja]
        col_a_fila2 = str(ws["A2"].value or "")
        if "EJEMPLO" not in col_a_fila2 and "EJEMPLO" not in str(ws["B2"].value or ""):
            errores.append(f"{hoja}: la fila 2 no parece ser un EJEMPLO")

    return errores


if __name__ == "__main__":
    errores = verificar()
    if errores:
        print("FALLÓ la verificación:")
        for e in errores:
            print(f"  - {e}")
        raise SystemExit(1)
    print("OK — Excel_de_Ventas_CEA.xlsx pasa la verificación estructural.")
